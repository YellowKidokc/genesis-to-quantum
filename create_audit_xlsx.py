import csv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict

csv_path = r'D:\GitHub\genesis-to-quantum\html_audit.csv'
output_path = r'D:\GitHub\genesis-to-quantum\HTML_AUDIT.xlsx'

# Read CSV
with open(csv_path, 'r', encoding='utf-8') as f:
    reader = csv.DictReader(f)
    rows = list(reader)

# Create workbook
wb = Workbook()
wb.remove(wb.active)

# Define styles
header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
header_font = Font(bold=True, color='FFFFFF', size=11)
linked_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
orphaned_fill = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')
category_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
category_font = Font(bold=True, size=10)
center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_align = Alignment(horizontal='left', vertical='top', wrap_text=True)
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

def style_header_row(sheet, row_num):
    for col in range(1, 6):
        cell = sheet.cell(row=row_num, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border

def add_data_rows(sheet, data_rows, start_row=2):
    for row_idx, row_data in enumerate(data_rows, start=start_row):
        sheet.cell(row=row_idx, column=1, value=row_data['filename'])
        sheet.cell(row=row_idx, column=2, value=row_data['title'])
        sheet.cell(row=row_idx, column=3, value=row_data['category'])
        sheet.cell(row=row_idx, column=4, value=row_data['is_linked'])
        sheet.cell(row=row_idx, column=5, value=row_data['linked_from'])
        
        # Apply fill based on linked status
        fill_color = linked_fill if row_data['is_linked'] == 'YES' else orphaned_fill
        for col in range(1, 6):
            cell = sheet.cell(row=row_idx, column=col)
            cell.fill = fill_color
            cell.border = border
            if col == 1:
                cell.alignment = left_align
            elif col == 4:
                cell.alignment = center_align
            else:
                cell.alignment = left_align

# SHEET 1: Summary & Overview
summary = wb.create_sheet('SUMMARY', 0)
linked = [r for r in rows if r['is_linked'] == 'YES']
orphaned = [r for r in rows if r['is_linked'] == 'NO']
by_cat = defaultdict(list)
for r in rows:
    by_cat[r['category']].append(r)

summary['A1'] = 'HTML AUDIT SUMMARY'
summary['A1'].font = Font(bold=True, size=14)
summary.merge_cells('A1:B1')

summary['A3'] = 'Total HTML Files:'
summary['B3'] = len(rows)
summary['A4'] = 'Linked Files (in navigation):'
summary['B4'] = len(linked)
summary['A5'] = 'Orphaned Files (not linked):'
summary['B5'] = len(orphaned)
summary['A6'] = 'Orphaned Percentage:'
summary['B6'] = f'{100*len(orphaned)//len(rows)}%'

for row in range(3, 7):
    summary[f'A{row}'].font = Font(bold=True)
    summary[f'B{row}'].border = border

summary['A9'] = 'By Category'
summary['A9'].font = Font(bold=True, size=11)
summary['B9'] = 'Total'
summary['C9'] = 'Orphaned'
for col in ['A', 'B', 'C']:
    summary[f'{col}9'].fill = category_fill
    summary[f'{col}9'].font = category_font
    summary[f'{col}9'].border = border

cat_row = 10
for cat in sorted(by_cat.keys()):
    cat_rows = by_cat[cat]
    cat_orphaned = sum(1 for r in cat_rows if r['is_linked'] == 'NO')
    summary[f'A{cat_row}'] = cat
    summary[f'B{cat_row}'] = len(cat_rows)
    summary[f'C{cat_row}'] = cat_orphaned
    for col in ['A', 'B', 'C']:
        summary[f'{col}{cat_row}'].border = border
    cat_row += 1

summary.column_dimensions['A'].width = 35
summary.column_dimensions['B'].width = 15
summary.column_dimensions['C'].width = 15

# SHEET 2: All Files (with color coding)
all_sheet = wb.create_sheet('ALL FILES', 1)
all_sheet['A1'] = 'Filename'
all_sheet['B1'] = 'Title'
all_sheet['C1'] = 'Category'
all_sheet['D1'] = 'Linked'
all_sheet['E1'] = 'Linked From'
style_header_row(all_sheet, 1)

add_data_rows(all_sheet, sorted(rows, key=lambda x: (x['category'], x['filename'])))

all_sheet.column_dimensions['A'].width = 50
all_sheet.column_dimensions['B'].width = 45
all_sheet.column_dimensions['C'].width = 25
all_sheet.column_dimensions['D'].width = 10
all_sheet.column_dimensions['E'].width = 30

# SHEET 3: Orphaned Files
orphaned_sheet = wb.create_sheet('ORPHANED (23)', 2)
orphaned_sheet['A1'] = 'Filename'
orphaned_sheet['B1'] = 'Title'
orphaned_sheet['C1'] = 'Category'
orphaned_sheet['D1'] = 'Status'
orphaned_sheet['E1'] = 'Note'
style_header_row(orphaned_sheet, 1)

add_data_rows(orphaned_sheet, sorted(orphaned, key=lambda x: (x['category'], x['filename'])))

orphaned_sheet.column_dimensions['A'].width = 50
orphaned_sheet.column_dimensions['B'].width = 45
orphaned_sheet.column_dimensions['C'].width = 25
orphaned_sheet.column_dimensions['D'].width = 10
orphaned_sheet.column_dimensions['E'].width = 30

# SHEET 4: Linked Files
linked_sheet = wb.create_sheet('LINKED (52)', 3)
linked_sheet['A1'] = 'Filename'
linked_sheet['B1'] = 'Title'
linked_sheet['C1'] = 'Category'
linked_sheet['D1'] = 'Status'
linked_sheet['E1'] = 'Linked From'
style_header_row(linked_sheet, 1)

add_data_rows(linked_sheet, sorted(linked, key=lambda x: (x['category'], x['filename'])))

linked_sheet.column_dimensions['A'].width = 50
linked_sheet.column_dimensions['B'].width = 45
linked_sheet.column_dimensions['C'].width = 25
linked_sheet.column_dimensions['D'].width = 10
linked_sheet.column_dimensions['E'].width = 30

# Save
wb.save(output_path)
print('Excel audit saved to ' + output_path)
print('   - Summary sheet with category breakdown')
print('   - All Files (' + str(len(rows)) + ' total)')
print('   - Orphaned Files (' + str(len(orphaned)) + ' files)')
print('   - Linked Files (' + str(len(linked)) + ' files)')
